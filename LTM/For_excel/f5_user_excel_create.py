import requests
import json
from requests.packages import urllib3
from nacl.pwhash import verify
import xlrd
import openpyxl
import pandas as pd
from xpinyin import Pinyin
from idlelib.iomenu import encoding
import getpass

p = Pinyin()
urllib3.disable_warnings()   #disable ssl warning

#導入excel數據    ------begin----------
config_file = input("請輸入需要導入的excel表格名稱，請帶尾碼如：配置資訊.xlsx，設定檔請放在python腳本一個資料夾下:")
data = xlrd.open_workbook(config_file,'r')
#data = pd.read_excel(config_file, sheet_name="sheet1", engine='openpyxl')
#data = pd.read_excel(config_file, sheet_name=0, engine='openpyxl')
#df=pandas.read_excel(‘data.xlsx’,engine=‘openpyxl’)

table = data.sheets()[0]
nrows = table.nrows
ncols = table.ncols

#導入excel數據  --------end-----------

#create auth token
f5_ip = input("請輸入F5管理IP:")
username = input("請輸入F5管理帳戶:")
password = getpass.getpass("請輸入F5管理密碼:")

req_url = 'https://%s/mgmt/shared/authn/login' %(f5_ip)
data = {'username':username,
        'password':password,
        'loginProviderName':'tmos'
        }
r = requests.post(req_url,json.dumps(data),verify = False)

#get auth token
f5_rep = json.loads(r.text)
f5_token = f5_rep['token']['token'] 

#header for connection f5
creat_header = {
    'X-F5-Auth-Token':f5_token,
    'Content-Type':'application/json'
    }

#create new pool
for l in range(1,nrows):
    if table.cell(l,0).ctype != 0:
        pool_url = 'https://%s/mgmt/tm/ltm/pool/' %(f5_ip)
        desc = p.get_pinyin(table.cell(l,0).value)
        name_to_py = p.get_pinyin("%s" %(table.cell(l,0).value),"")
        pool_name = '_1_%s_pool' %(name_to_py)
        print("創建的pool:",pool_name)
        creat_pool_data = {
            'partition':'Common',
            'name':pool_name,
            'description':desc,
            'monitor':'http'
            }
        pool = requests.post(url = pool_url,headers = creat_header, data = json.dumps(creat_pool_data),verify = False)
    else :
        continue
#members into pool
for l in range(1,nrows):
    pool_name_last = pool_name
    name_to_py = p.get_pinyin("%s" %(table.cell(l,0).value),"")
    if table.cell(l,0).ctype == 0 :
        pool_name = pool_name_last
    else :
        pool_name = '_1_%s_pool' %(name_to_py)
    pool_member_ip = table.cell(l,4).value
    pool_member_port = table.cell(l,5).value
    pool_member_url = 'https://%s/mgmt/tm/ltm/pool/%s/members' %(f5_ip,pool_name)
    member_into_pool_data = {
        'partition':'Common',
        'name':'%s:%s' % (pool_member_ip,pool_member_port),'address':pool_member_ip
        }
    pool = requests.post(url = pool_member_url,headers = creat_header, data = json.dumps(member_into_pool_data),verify = False)

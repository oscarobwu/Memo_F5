#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       1234.py
#
#        USAGE: 1234.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 16:39:19
#      Last modified: 2021-01-06 16:49
#     REVISION: ---
#===============================================================================
from openpyxl import Workbook
import openpyxl

file = "2021-01-06-16-04_vs_pool.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
ws = wb.active

for row in ws.iter_rows("F"):
    for cell in row:
        if cell.value == str("pool_8086_172.99.1.100"):
            print(ws.cell(row=cell.row, column=2).value) #change column number for any cell value you want

#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       print_execl_column1.py
#
#        USAGE: print_execl_column1.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 20:22:22
#      Last modified: 2021-01-06 20:29
#     REVISION: ---
#===============================================================================
from openpyxl import load_workbook
def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f'Worksheet names: {workbook.sheetnames}')
    sheet = workbook.active
    print(sheet)
    print(f'The title of the Worksheet is: {sheet.title}')
if __name__ == '__main__':
    open_workbook('2021-01-06-16-04_vs_pool.xlsx')

# set file path
filepath="2021-01-06-16-04_vs_pool.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# get b1 cell value
b1=sheet['B1']
# get b2 cell value
b2=sheet['C1']
# get b3 cell value
b3=sheet.cell(row=3,column=2)
# print b1, b2 and b3
print(b1.value)
print(b2.value)
print(b3.value)

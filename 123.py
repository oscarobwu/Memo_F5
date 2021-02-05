#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       123.py
#
#        USAGE: 123.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 15:56:56
#      Last modified: 2021-01-06 16:00
#     REVISION: ---
#===============================================================================
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
#
wb = Workbook()
ws = wb.active
ws['B3'] = "Hello"
ws['B3'].font = Font(color="ff0000", bold=True)
wb.save("BoldDemo.xlsx")

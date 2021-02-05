#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       test.py
#
#        USAGE: test.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-05 20:14:16
#      Last modified: 2021-01-05 20:14
#     REVISION: ---
#===============================================================================
import openpyxl
wb = openpyxl.Workbook()
#ws1 = wb.active
ws1 = wb.get_sheet_by_name('Sheet')
ws1.title = "Test"
ws2 = wb.create_sheet("ABC")
wb.create_sheet(u"中文表名")
ws1.sheet_properties.tabColor="0000FF"
ws2.sheet_properties.tabColor="FF00FF"
ws1['A1'] = "哇嘎嘎"
ws1['B2'] = "嗯~ o(*￣▽￣*)o"
wb.save("test.xlsx")

#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       read_first_column.py
#
#        USAGE: read_first_column.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-07 13:59:16
#      Last modified: 2021-01-07 14:24
#     REVISION: ---
#===============================================================================
import openpyxl
#from openpyxl.worksheet import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Give the location of the file
path = "2021-01-06-16-04_vs_pool.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

#Loop will print all columns name
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    column = cell_obj.coordinate
    xx = cell_obj.coordinate
    #print(cell_obj.coordinate)
    ww = xx[0]
    yy = cell_obj.column
    print("column: {} , xx : {} , yy : {} , 顯示欄位英文名稱: : {}".format(column, xx, yy, ww))

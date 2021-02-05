#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       1237.py
#
#        USAGE: 1237.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 17:30:12
#      Last modified: 2021-01-06 20:03
#     REVISION: ---
#===============================================================================
from openpyxl.workbook import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
#
#Path
wb = openpyxl.load_workbook(r'2021-01-06-16-04_vs_pool.xlsx')

#active worksheet data
ws = wb.active

searchString = input('\nFind searchString: ')

for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        if searchString == ws.cell(i,j).value:
            print("found")
            print(ws.cell(i,j))
            sss = ws.cell(i,j).coordinate
            #print(sss)
            ws[sss].font = Font(color="ff0000", bold=True)
            ws[sss].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
            #for x in sss:
            #    print(x)
            #    ws[x].font = Font(color="ff0000", bold=True)

wb.save('2021-01-06-16-04_vs_pool.xlsx')

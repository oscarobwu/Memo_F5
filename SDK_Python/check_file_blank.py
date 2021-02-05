#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       check_file_blank.py
#
#        USAGE: check_file_blank.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-02-02 12:02:19
#      Last modified: 2021-02-02 14:08
#     REVISION: ---
#===============================================================================
from openpyxl import load_workbook
File_Name = input('File_name: ')  #要处理的文件路径
file = File_Name
wb = load_workbook(file)  #加载文件
ws = wb.active
for i in range(1,ws.max_row+1):
    for j in range(1,ws.max_column+1):
        old = ws.cell(i, j).value.str
        if old is not None:
            #ws.cell(i, j).value = old.strip().replace(' ', '').replace("\n", "")
            ws.cell(i, j).value = old.str.strip().replace(' ', '').replace("\n", "")
wb.save(file)
wb.close()
print("處理完成")

#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       1236.py
#
#        USAGE: 1236.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 17:20:51
#      Last modified: 2021-01-06 17:21
#     REVISION: ---
#===============================================================================
###################################
import openpyxl
#Path
wb = openpyxl.load_workbook(r'2021-01-06-16-04_vs_pool.xlsx')

#active worksheet data
ws = wb.active

searchString = input('\nFind searchString: ')

for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        if searchString == ws.cell(i,j).value:
            print("found")
            print(ws.cell(i,j))

#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       1233.py
#
#        USAGE: 1233.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 16:00:41
#      Last modified: 2021-01-06 16:28
#     REVISION: ---
#===============================================================================
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import openpyxl
#
ref_workbook = openpyxl.load_workbook('2021-01-06-16-04_vs_pool.xlsx')
#
#wb = Workbook()
ws = ref_workbook.active
#ws['B3'] = "Hello"
#cell_range = ws['B3':'B5']
#cell_range.font = Font(color="ff0000", bold=True)
lista = ['B3', 'B4', 'B5']
for x in lista:
    ws[x].font = Font(color="ff0000", bold=True)
#wb.save("BoldDemo.xlsx")
ref_workbook.save('2021-01-06-16-04_vs_pool.xlsx')
#

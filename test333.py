#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       test333.py
#
#        USAGE: test333.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-07 18:02:22
#      Last modified: 2021-01-07 18:02
#     REVISION: ---
#===============================================================================
import pandas as pd
import openpyxl
import datetime
import time
from f5.bigip import ManagementRoot
import datetime
import time
import getpass
import operator
from openpyxl.workbook import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
#
f5_host = input('F5_mgmt_IP: ')
#f5_host = '192.168.8.168'
print("登入 主機 : %s" % f5_host)
f5_user = input('\nUsername: ')
f5_pw = getpass.getpass('\nPassword: ')
#
#findword = input('\n需要搜尋比對字串: ')
#
#print("登入 主機 : %s" % f5_host)
datestring = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M");
#mgmt = ManagementRoot('hostname', 'username', 'password')
mgmt = ManagementRoot(f5_host, f5_user, f5_pw)

# CSV header
#print('Partition, VS Name, Pool Name, Pool Members')

#df_list = pd.DataFrame()
df_list = pd.DataFrame(columns=['vs_part',
                                'vs_name',
                                'vs_ip_addr',
                                'vs_ip_port',
                                'vs_pool',
                                'Pool_memb',
                                'Pool_memb_addr'])

for virtual in mgmt.tm.ltm.virtuals.get_collection():
    # Does the virtual server have a pool assigned?
    if not getattr(virtual, 'pool', None):
        # No - print virtual server name and move on
        vs_ip_addr = virtual.destination.split('/')[2]
        vs_ip_port = virtual.destination.split(':')[1]
        #print('{},{},,'.format(virtual.partition, virtual.name))
        df_list = df_list.append({'vs_part': virtual.partition,
                                  'vs_name': virtual.name,
                                  'vs_ip_addr': vs_ip_addr,
                                  'vs_ip_port': vs_ip_port,
                                  },
                     ignore_index=True)
        continue

    # Pool partition
    vs_ip_addr = virtual.destination.split('/')[2]
    vs_ip_port = virtual.destination.split(':')[1]
    #print(virtual.destination)
    #print(vs_ip_addr)
    #print(vs_ip_port)
    # Pool partition
    pool_part = virtual.pool.split('/')[1]
    # Pool name
    pool_name = virtual.pool.split('/')[2]

    pool = mgmt.tm.ltm.pools.pool.load(
        partition=pool_part,
        name=pool_name,
    )

    members = pool.members_s.get_collection()

    # Loop to output the results however you'd like
    # Example is CSV with pool members delimited by semicolon

    # Gather members in list - makes printing easier
    pool_members = []
    pool_members_addr = []
    for member in members:
        pool_members.append(member.name)
        pool_members_addr.append(member.address)

    df_list = df_list.append({'vs_part': virtual.partition,
                              'vs_name': virtual.name,
                              'vs_ip_addr': vs_ip_addr,
                              'vs_ip_port': vs_ip_port,
                              'vs_pool': pool.name,
                              'Pool_memb': '{}'.format('\r\n'.join(pool_members)),
                              'Pool_memb_addr': '{}'.format('\r\n'.join(pool_members_addr))},
                 ignore_index=True)

print("等待最後輸出\n")
time.sleep(3)
print(df_list)
#from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
#
ws = wb.active
rows = dataframe_to_rows(df_list)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
        colist =['G', 'H']
        for x in colist:
            ws.column_dimensions[x].width = 30.0
            #print("set column width {}".format(x))
            for cell in ws[x]:
                cell.alignment = Alignment(wrapText=True)
                #print(cell.value)


searchString = input('\n需要搜尋比對字串 - Find searchString: ')
#
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        if searchString == ws.cell(i,j).value:
            print("found")
            print(ws.cell(i,j))
            sss = ws.cell(i,j).coordinate
            #print(sss)
            ws[sss].font = Font(color="ff0000", bold=True)
            ws[sss].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
            #    ws[x].font = Font(color="ff0000", bold=True)

ws.title = f5_host
ws.sheet_properties.tabColor = "1072BA"
wb.save(filename = datestring + '_vs_pool.xlsx')


#!/usr/bin/python
#-*- coding: utf-8 -*-
#===============================================================================
#
#         Filename:       print_execl_column.py
#
#        USAGE: print_execl_column.py
#
#  DESCRIPTION: 
#
#      OPTIONS: ---
# REQUIREMENTS: ---
#         BUGS: ---
#        NOTES: ---
#       AUTHOR: Oscarob Wu(oscarobwu@gmail.com), 
# ORGANIZATION: 
#      VERSION: 1.0
#      Created Time: 2021-01-06 20:15:05
#      Last modified: 2021-01-06 20:15
#     REVISION: ---
#===============================================================================
import openpyxl

wb = openpyxl.load_workbook(filename='2021-01-06-16-04_vs_pool.xlsx', use_iterators=True)
ws = wb.get_sheet_by_name('Pool_memb')

for index, row in enumerate(ws.iter_rows()):
    for cell in row:
        print(ws.cell(row=index + 1, column=1).value, cell.value)

